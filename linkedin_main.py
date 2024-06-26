import time
import pickle
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException, ElementClickInterceptedException
from openpyxl import Workbook, load_workbook

def save_cookies(driver, filepath):
    with open(filepath, 'wb') as file:
        pickle.dump(driver.get_cookies(), file)

def load_cookies(driver, filepath):
    with open(filepath, 'rb') as file:
        cookies = pickle.load(file)
        for cookie in cookies:
            driver.add_cookie(cookie)

def login_linkedin(driver, username=None, password=None):
    driver.get('https://www.linkedin.com/login')

    # Load cookies if available
    try:
        load_cookies(driver, "cookies.pkl")
        driver.refresh()
        return True  # Return True if login successful using cookies
    except (FileNotFoundError, IOError, pickle.UnpicklingError):
        pass

    # If not logged in or cookies not available, proceed with login
    if username and password:
        username_input = driver.find_element(By.ID, 'username')
        password_input = driver.find_element(By.ID, 'password')

        username_input.send_keys(username)
        password_input.send_keys(password)
        driver.find_element(By.XPATH, '//button[@type="submit"]').click()

        # Perform 2FA if required
        input("Press Enter after completing 2FA...")

        # Save cookies for future sessions
        save_cookies(driver, "cookies.pkl")

        return True

    return False

def navigate_to_company_pages(driver):
    driver.get('https://www.linkedin.com/mynetwork/network-manager/company/')
    time.sleep(5)  # Give the page time to load initially

def scroll_to_load_more_companies(driver):
    last_height = driver.execute_script("return document.body.scrollHeight")
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(5)  # Wait for new companies to load
    new_height = driver.execute_script("return document.body.scrollHeight")
    return new_height != last_height

def like_posts_and_scrape_company_info(driver, company_index, visited_companies, wb):
    try:
        companies = driver.find_elements(By.CSS_SELECTOR, '.reusable-search__result-container')
        company = companies[company_index]
        company_name = company.text.split('\n')[0]  # Extract the company name

        if company_name in visited_companies:
            print(f"Skipping already visited company: {company_name}")
            return

        company.click()

        wait = WebDriverWait(driver, 20)
        try:
            posts_tab = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//a[contains(@class, "org-page-navigation__item-anchor") and contains(@href, "/posts/")]')
            ))
            posts_tab.click()
            time.sleep(10)
        except TimeoutException:
            print(f"Timeout while waiting for posts tab for company {company_name}.")
            driver.get('https://www.linkedin.com/mynetwork/network-manager/company/')
            return

        posts_liked = 0
        last_height = driver.execute_script("return document.body.scrollHeight")
        
        while posts_liked < 50:
            try:
                posts = driver.find_elements(By.CSS_SELECTOR, 'button[aria-label="React Like"]')
                for post in posts:
                    if 'Like' in post.text and posts_liked < 50:
                        try:
                            post.click()
                            posts_liked += 1
                            time.sleep(2)
                        except ElementClickInterceptedException:
                            driver.execute_script("arguments[0].scrollIntoView(true);", post)
                            time.sleep(2)
                            driver.execute_script("window.scrollBy(0, -150);")
                            time.sleep(2)
                            post.click()
                            posts_liked += 1
                            time.sleep(2)

                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(5)
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
            except NoSuchElementException:
                print("No posts found or unable to interact with posts.")
                break
            except StaleElementReferenceException:
                print("Stale element exception. Retrying...")
                continue

        try:
            about_tab = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//a[contains(@class, "org-page-navigation__item-anchor") and contains(@href, "/about/")]')
            ))
            about_tab.click()
            time.sleep(5)

            name = "N/A"
            website = "N/A"
            overview = "N/A"
            size = "N/A"
            headquarters = "N/A"
            phone = "N/A"

            # Attempt to scrape each element
            try:
                name = driver.find_element(By.XPATH, '//h1[contains(@class, "org-top-card-summary__title")]').text
            except NoSuchElementException:
                pass

            try:
                website = driver.find_element(By.XPATH, '//dd[@class="mb4 t-black--light text-body-medium"]/a').get_attribute('href')
            except NoSuchElementException:
                pass

            try:
                overview = driver.find_element(By.XPATH, '//h2[text()="Overview"]/following-sibling::p').text
            except NoSuchElementException:
                pass

            try:
                size = driver.find_element(By.XPATH, '//dt[text()="Company size"]/following-sibling::dd').text
            except NoSuchElementException:
                pass

            try:
                headquarters = driver.find_element(By.XPATH, '//dt[text()="Headquarters"]/following-sibling::dd').text
            except NoSuchElementException:
                pass
            
            try:
                phone_element = driver.find_element(By.XPATH, '//a[contains(@href, "tel:")]/span[contains(@class, "link-without-visited-state")]')
                phone = phone_element.text.strip() if phone_element else "N/A"
            except NoSuchElementException:
                pass

            company_info = {
                'name': name,
                'website': website,
                'overview': overview,
                'size': size,
                'headquarters': headquarters,
                'phone': phone
            }

            print(company_info)

            visited_companies.add(company_name)

            append_to_excel(wb, company_info)

        except TimeoutException:
            print("Timeout waiting for the 'About' tab to be clickable.")
        except NoSuchElementException as e:
            print(f"Error scraping company info: {e}")
        except Exception as e:
            print(f"Unexpected error scraping company info: {e}")

        # Save the workbook immediately after visiting each company
        wb.save("visited_companies.xlsx")
        print("Excel file saved.")

        save_visited_companies(visited_companies)
        print("Visited company list is updated.")

        driver.get('https://www.linkedin.com/mynetwork/network-manager/company/')
    except IndexError:
        print("No more companies to process.")
        return

def save_visited_companies(visited_companies):
    with open("visited_companies.pkl", 'wb') as file:
        pickle.dump(visited_companies, file)

def load_visited_companies():
    try:
        with open("visited_companies.pkl", 'rb') as file:
            visited_companies = pickle.load(file)
    except (FileNotFoundError, IOError, pickle.UnpicklingError):
        visited_companies = set()
    return visited_companies

def append_to_excel(wb, company_info):
    ws = wb.active
    row = (
        company_info['name'],
        company_info['website'],
        company_info['overview'],
        company_info['size'],
        company_info['headquarters'],
        company_info['phone']
    )
    ws.append(row)

def main():
    # Check if cookies are available
    cookies_available = False
    try:
        with open("cookies.pkl", 'rb') as file:
            cookies = pickle.load(file)
            if cookies:
                cookies_available = True
    except (FileNotFoundError, IOError, pickle.UnpicklingError):
        pass

    if not cookies_available:
        # If cookies are not available, ask for username and password
        username = input("Enter your LinkedIn username: ")
        password = input("Enter your LinkedIn password: ")
    else:
        username = None
        password = None

    # Specify the path to the ChromeDriver in the project directory
    driver_path = 'chromedriver-win64/chromedriver.exe'
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")

    chrome_service = ChromeService(executable_path=driver_path)

    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

    # Login to LinkedIn
    login_success = login_linkedin(driver, username, password)
    if not login_success:
        print("Login failed. Exiting script.")
        driver.quit()
        return

    # Load or create the workbook
    try:
        wb = load_workbook("visited_companies.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Company Name", "Website", "Overview", "Size", "Headquarters", "Phone"])
    else:
        ws = wb.active

    visited_companies = load_visited_companies()
    print("Current visited companies:", [name.split('\n')[0] for name in visited_companies])

    navigate_to_company_pages(driver)

    company_index = 0
    start_time = time.time()  # Record the start time of the script

    while True:
        # Check if an hour has passed
        elapsed_time = time.time() - start_time
        if elapsed_time >= 3600:
            print("Pausing script for 10 minutes.")
            time.sleep(600)  # Pause for 10 minutes (600 seconds)
            start_time = time.time()  # Reset the start time

        # Get the list of companies
        companies = driver.find_elements(By.CSS_SELECTOR, '.reusable-search__result-container')

        # Skip already visited companies
        while company_index < len(companies) and companies[company_index].text.split('\n')[0] in visited_companies:
            company_index += 1

        if company_index >= len(companies):
            print("No more new companies to process. Scrolling to load more companies.")
            if not scroll_to_load_more_companies(driver):
                print("No more companies to process even after scrolling. Exiting.")
                break
            companies = driver.find_elements(By.CSS_SELECTOR, '.reusable-search__result-container')
            company_index = 0
            continue

        like_posts_and_scrape_company_info(driver, company_index, visited_companies, wb)
        company_index += 1
        time.sleep(20)  # Adjust sleep time as necessary

        save_visited_companies(visited_companies)
        print("Visited company list is updated.")

    driver.quit()

if __name__ == "__main__":
    main()
